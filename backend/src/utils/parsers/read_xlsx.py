from datetime import datetime

from openpyxl import load_workbook

def read_xlsx(file, dates=None, floats=None):
    '''Reads an xlsx file to a list of tuples
    The first row of the spreadsheet is assumed to contain headers
    The floats argument gives a list of column names that are of the float datatype
    The dates argument gives a list of column names that are of the datetime datatype
        dates are assumed to be specified in the format d/m/yyyy
    '''
    wb = load_workbook(file)
    ws = wb.active
    headers = next(ws.values)
    float_columns = None
    if floats:
        float_columns = [i for i, head in enumerate(headers) if head in floats]
    date_columns = None
    if dates:
        date_columns = [i for i, head in enumerate(headers) if head in dates]
    result = [tuple(headers)]
    for row in ws.iter_rows(min_row=2, values_only=True):
        parsed_row = []
        for i, e in enumerate(row):
            if i in float_columns:
                parsed_row.append(float(e))
            elif i in date_columns:
                parsed_row.append(datetime.strptime(e, '%m/%d/%Y'))
            else:
                parsed_row.append(e)
        result.append(parsed_row)
    return result


# if __name__ == '__main__':
#     # test
#     file = '../../db/data/dev-data.xlsx'
#     data = read_xlsx(file, dates=['date'], floats=['amount'])
#     print(data)
